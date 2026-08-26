"""Generates a synthetic SF1 EXPORT/generation template (.xlsx) for
Wave 3's official-form engine.

SYNTHETIC TEST DATA ONLY -- this is NOT an official DepEd SF1 template.
No authoritative SF1 export template was available in this repository or
this development environment (see docs/adr/0048-official-form-engine-sf1.md's
"Authoritative-template evidence gate" section). This fixture exists
solely to prove the engine's template-copy/cell-patching/fidelity
mechanics against a template that deliberately exercises the structural
features the engine must preserve: a merged title, a merged header-info
block, a formula OUTSIDE the write range, a second untouched+protected
sheet, custom row heights/column widths, borders, and a defined print
area.

Distinct from tests/fixtures/generate_sf1_fixtures.py, which builds
*import*-direction (.xls) fixtures for reading a teacher-provided
workbook -- this script builds an *export/generation*-direction (.xlsx)
fixture: a template the engine fills in.

Re-run this script only if the fixture's structure itself needs to
change; do not hand-edit the generated .xlsx file (openpyxl-authored
files are not reliably hand-editable without corrupting their internal
structure).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.dimensions import ColumnDimension

MAX_LEARNER_ROWS = 30
FIRST_DATA_ROW = 9  # 1-based; header labels occupy rows 1-8

def build():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "SF1"

    # --- Title (merged) ---
    sheet.merge_cells("A1:D1")
    sheet["A1"] = (
        "SCHOOL FORM 1 (SF1) SCHOOL REGISTER "
        "-- SYNTHETIC TEMPLATE, NOT AN OFFICIAL DEPED DOCUMENT"
    )
    sheet["A1"].font = Font(bold=True, size=12)
    sheet["A1"].alignment = Alignment(horizontal="center")

    # --- Header info block (merged label targets) ---
    sheet["A3"] = "School Name:"
    sheet.merge_cells("B3:D3")
    sheet["A4"] = "School Year:"
    sheet.merge_cells("B4:D4")
    sheet["A5"] = "Grade Level:"
    sheet.merge_cells("B5:D5")
    sheet["A6"] = "Section:"
    sheet.merge_cells("B6:D6")

    # --- Column headers (row 8) ---
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    for col, label in zip("ABCD", ["LRN", "Family Name", "Given Name", "Sex"]):
        cell = sheet[f"{col}8"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border

    # --- Reserved learner data rows (left blank; the engine fills these) ---
    last_data_row = FIRST_DATA_ROW + MAX_LEARNER_ROWS - 1  # row 38
    for row in range(FIRST_DATA_ROW, last_data_row + 1):
        for col in "ABCD":
            sheet[f"{col}{row}"].border = border

    # --- Footer formula OUTSIDE the write range -- must survive untouched ---
    footer_row = last_data_row + 2  # row 40
    sheet[f"A{footer_row}"] = "Total Learners:"
    sheet[f"A{footer_row}"].font = Font(bold=True)
    sheet[f"B{footer_row}"] = f"=COUNTA(A{FIRST_DATA_ROW}:A{last_data_row})"

    # --- Sizing / layout the engine must not disturb ---
    sheet.row_dimensions[1].height = 22
    sheet.row_dimensions[8].height = 18
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 20
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 10
    sheet.freeze_panes = "A9"
    sheet.print_area = f"A1:D{footer_row}"

    # --- Second sheet: untouched by generation, sheet-protected ---
    notes = wb.create_sheet("Notes")
    notes["A1"] = (
        "SYNTHETIC FIXTURE -- for LIKHA-SIS Wave 3 official-form engine "
        "testing only. Not an official DepEd document. This sheet must "
        "remain unchanged by SF1 generation."
    )
    notes.protection.sheet = True
    notes.protection.password = "synthetic-test-only"

    wb.save("sf1_template_synthetic.xlsx")


if __name__ == "__main__":
    build()
    print("wrote sf1_template_synthetic.xlsx")
